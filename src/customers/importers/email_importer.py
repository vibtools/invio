from __future__ import annotations

import csv
import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from ...core.dynamic_tags import contains_supported_dynamic_tag
from ..models import CustomerRecord

EMAIL_RE = re.compile(r"[A-Z0-9.!#$%&'*+/=?^_`{|}~-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)
EMAIL_FULL_RE = re.compile(r"^[A-Z0-9.!#$%&'*+/=?^_`{|}~-]+@[A-Z0-9.-]+\.[A-Z]{2,}$", re.I)
STRUCTURED_HEADERS = {"email", "name", "country"}


@dataclass(frozen=True, slots=True)
class CustomerImportIssue:
    row_number: int
    message: str
    kind: str = "invalid"

    def display(self) -> str:
        if self.row_number > 0:
            return f"Row {self.row_number}: {self.message}"
        return self.message


@dataclass(slots=True)
class CustomerImportResult:
    records: list[CustomerRecord] = field(default_factory=list)
    record_rows: list[int] = field(default_factory=list)
    issues: list[CustomerImportIssue] = field(default_factory=list)
    duplicates_skipped: int = 0
    structured: bool = False

    @property
    def warnings(self) -> list[str]:
        return [issue.display() for issue in self.issues]


def _collect(values: list[object]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for value in values:
        for match in EMAIL_RE.findall(str(value or "")):
            email = match.strip().lower()
            if email not in seen:
                seen.add(email)
                result.append(email)
    return result


def _read_rows(file_path: Path) -> list[tuple[int, list[object]]]:
    suffix = file_path.suffix.lower()
    rows: list[tuple[int, list[object]]] = []
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), 1):
                rows.append((row_number, list(row)))
        finally:
            workbook.close()
        return rows
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row_number, row in enumerate(csv.reader(handle, delimiter=delimiter), 1):
                rows.append((row_number, list(row)))
        return rows
    if suffix == ".txt":
        with file_path.open("r", encoding="utf-8-sig", errors="replace") as handle:
            for row_number, line in enumerate(handle, 1):
                rows.append((row_number, [line.rstrip("\r\n")]))
        return rows
    raise ValueError("Supported customer files are CSV, TSV, XLSX, XLSM and TXT.")


def _first_usable_row(rows: list[tuple[int, list[object]]]) -> tuple[int, list[object]] | None:
    for row_number, values in rows:
        if any(str(value or "").strip() for value in values):
            return row_number, values
    return None


def _header_map(values: list[object]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(values):
        header = str(value or "").strip().casefold()
        if header in STRUCTURED_HEADERS and header not in result:
            result[header] = index
    return result


def _structured_import(rows: list[tuple[int, list[object]]], header_row: int, headers: dict[str, int]) -> CustomerImportResult:
    result = CustomerImportResult(structured=True)
    accepted_by_email: dict[str, CustomerRecord] = {}

    def value_at(values: list[object], name: str) -> str:
        index = headers.get(name)
        if index is None or index >= len(values):
            return ""
        return str(values[index] or "").strip()

    for row_number, values in rows:
        if row_number <= header_row:
            continue
        if not any(str(value or "").strip() for value in values):
            continue
        email_raw = value_at(values, "email")
        if not email_raw:
            result.issues.append(CustomerImportIssue(row_number, "email is missing."))
            continue
        email = email_raw.lower()
        if EMAIL_FULL_RE.fullmatch(email) is None:
            result.issues.append(CustomerImportIssue(row_number, "email format is invalid."))
            continue
        name = value_at(values, "name")
        country = value_at(values, "country").upper()
        if country and (len(country) != 2 or not country.isascii() or not country.isalpha()):
            result.issues.append(CustomerImportIssue(row_number, "country must be a two-letter code."))
            continue
        record = CustomerRecord(email=email, name=name, country=country)
        previous = accepted_by_email.get(record.email)
        if previous is not None:
            if previous.name == record.name and previous.country == record.country:
                result.duplicates_skipped += 1
            elif not record.name and not record.country:
                result.duplicates_skipped += 1
            else:
                result.issues.append(
                    CustomerImportIssue(
                        row_number,
                        f"duplicate customer data conflicts with the first row for {record.email}.",
                        kind="conflict",
                    )
                )
            continue
        accepted_by_email[record.email] = record
        result.records.append(record)
        result.record_rows.append(row_number)

    if not result.records and not result.issues:
        result.issues.append(CustomerImportIssue(0, "No customer rows were found."))
    return result


def _legacy_import(rows: list[tuple[int, list[object]]]) -> CustomerImportResult:
    result = CustomerImportResult(structured=False)
    seen: set[str] = set()
    for row_number, values in rows:
        for value in values:
            for match in EMAIL_RE.findall(str(value or "")):
                email = match.strip().lower()
                if email in seen:
                    result.duplicates_skipped += 1
                    continue
                seen.add(email)
                result.records.append(CustomerRecord(email=email))
                result.record_rows.append(row_number)
    if not result.records:
        result.issues.append(CustomerImportIssue(0, "No valid email addresses were found."))
    return result



def apply_customer_defaults(
    records: list[CustomerRecord],
    *,
    default_name: str = "",
    default_country: str = "",
) -> list[CustomerRecord]:
    """Apply owner-configured import defaults without mutating parsed records.

    An explicit configured default takes precedence. Otherwise existing structured
    name/country values are preserved; missing names use the email username and
    missing countries use ``US``.
    """

    configured_name = str(default_name).strip()
    configured_country = str(default_country).strip().upper()
    if configured_country and (
        len(configured_country) != 2 or not configured_country.isascii() or not configured_country.isalpha()
    ):
        raise ValueError("Default customer country must be a two-letter code when provided.")

    normalized: list[CustomerRecord] = []
    for record in records:
        email_username = record.email.split("@", 1)[0]
        name = configured_name or record.name or email_username
        country = configured_country or record.country or "US"
        normalized.append(
            CustomerRecord(
                record.email,
                name,
                country,
                name_is_dynamic=bool(configured_name and contains_supported_dynamic_tag(configured_name)),
            )
        )
    return normalized


def import_customers(path: str | Path) -> CustomerImportResult:
    """Import customer records with structured and legacy email-only compatibility.

    CSV/TSV/Excel files enter structured mode only when the first usable row
    contains an ``email`` header. TXT files and files without that header retain
    the historical email-extraction workflow. Name and country are never guessed.
    """

    file_path = Path(path)
    try:
        rows = _read_rows(file_path)
    except ValueError:
        raise
    except Exception as exc:
        # File parsing is a user-input boundary. Convert malformed workbook/CSV
        # parser failures into the existing ValueError contract so the UI can
        # report them without allowing an uncaught exception to escape.
        raise ValueError(f"Customer file could not be read: {exc}") from exc
    suffix = file_path.suffix.lower()
    if suffix in {".csv", ".tsv", ".xlsx", ".xlsm"}:
        first = _first_usable_row(rows)
        if first is not None:
            header_row, header_values = first
            headers = _header_map(header_values)
            if "email" in headers:
                return _structured_import(rows, header_row, headers)
    return _legacy_import(rows)


def import_emails(path: str | Path) -> tuple[list[str], list[str]]:
    """Extract unique emails using the original email-only compatibility contract."""

    file_path = Path(path)
    suffix = file_path.suffix.lower()
    values: list[object] = []

    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            for row in sheet.iter_rows(values_only=True):
                values.extend(row)
        finally:
            workbook.close()
    elif suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with file_path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.reader(handle, delimiter=delimiter):
                values.extend(row)
    else:
        values.append(file_path.read_text(encoding="utf-8-sig", errors="replace"))

    emails = _collect(values)
    warnings: list[str] = []
    if not emails:
        warnings.append("No valid email addresses were found.")
    return emails, warnings
